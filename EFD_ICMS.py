"""
EFD Extrator C100 + C170 + C190 — Entradas & Saídas
Extrai registros C100/C170 com NCM/CEST do 0200 (Entradas e Saídas)
e C190 consolidado (apenas Saídas), gerando XLSX formatado.
"""
# ─── AUTO-INSTALL ────────────────────────────────────────────────────────────
import subprocess, sys

def _install(pkg):
    try:
        __import__(pkg)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "--quiet"])

_install("openpyxl")

# ─── IMPORTS ─────────────────────────────────────────────────────────────────
import streamlit as st
import tempfile, zipfile, os, io, time
from pathlib import Path

# ─── LAYOUTS SPED ────────────────────────────────────────────────────────────
IDX_0200_COD_ITEM = 1
IDX_0200_COD_NCM  = 7
IDX_0200_CEST     = 12

C100_FIELDS = [
    "REG", "IND_OPER", "IND_EMIT", "COD_PART", "COD_MOD", "COD_SIT",
    "SER", "NUM_DOC", "CHV_NFE", "DT_DOC", "DT_E_S", "VL_DOC",
    "IND_PGTO", "VL_DESC", "VL_ABAT_NT", "VL_MERC", "IND_FRT",
    "VL_FRT", "VL_SEG", "VL_OUT_DA", "VL_BC_ICMS", "VL_ICMS",
    "VL_BC_ICMS_ST", "VL_ICMS_ST", "VL_IPI", "VL_PIS", "VL_COFINS",
    "VL_PIS_ST", "VL_COFINS_ST"
]

C170_FIELDS = [
    "REG", "NUM_ITEM", "COD_ITEM", "DESCR_COMPL", "QTD", "UNID",
    "VL_ITEM", "VL_DESC", "IND_MOV", "CST_ICMS", "CFOP", "COD_NAT",
    "VL_BC_ICMS", "ALIQ_ICMS", "VL_ICMS", "VL_BC_ICMS_ST", "ALIQ_ST",
    "VL_ICMS_ST", "IND_APUR", "CST_IPI", "COD_ENQ", "VL_BC_IPI",
    "ALIQ_IPI", "VL_IPI", "CST_PIS", "VL_BC_PIS", "ALIQ_PIS",
    "QUANT_BC_PIS", "ALIQ_PIS_QUANT", "VL_PIS", "CST_COFINS",
    "VL_BC_COFINS", "ALIQ_COFINS", "QUANT_BC_COFINS", "ALIQ_COFINS_QUANT",
    "VL_COFINS", "COD_CTA", "VL_ABAT_NT"
]

# Registro C190 — Consolidação por CST_ICMS, CFOP e ALIQ_ICMS
# |C190|CST_ICMS|CFOP|ALIQ_ICMS|VL_OPR|VL_BC_ICMS|VL_ICMS|VL_BC_ICMS_ST|VL_ICMS_ST|VL_RED_BC|VL_IPI|COD_OBS|
C190_FIELDS = [
    "REG", "CST_ICMS", "CFOP", "ALIQ_ICMS", "VL_OPR", "VL_BC_ICMS",
    "VL_ICMS", "VL_BC_ICMS_ST", "VL_ICMS_ST", "VL_RED_BC", "VL_IPI", "COD_OBS"
]

N_C100 = len(C100_FIELDS)
N_C170 = len(C170_FIELDS)
N_C190 = len(C190_FIELDS)

# ─── ÍNDICES PRÉ-CALCULADOS ─────────────────────────────────────────────────
_C100_IDX = {f: i for i, f in enumerate(C100_FIELDS)}
_C170_IDX = {f: i for i, f in enumerate(C170_FIELDS)}
_C190_IDX = {f: i for i, f in enumerate(C190_FIELDS)}

# ─── COLUNAS DE SAÍDA — ENTRADAS / SAÍDAS (C100+C170+0200) ──────────────────
OUTPUT_COLUMNS = [
    ("NUM_DOC",     "c100", _C100_IDX["NUM_DOC"]),
    ("COD_ITEM",    "c170", _C170_IDX["COD_ITEM"]),
    ("DESCR_COMPL", "c170", _C170_IDX["DESCR_COMPL"]),
    ("NCM",         "0200", 0),
    ("CEST",        "0200", 1),
    ("VL_ITEM",     "c170", _C170_IDX["VL_ITEM"]),
    ("CST_ICMS",    "c170", _C170_IDX["CST_ICMS"]),
    ("CFOP",        "c170", _C170_IDX["CFOP"]),
    ("VL_BC_ICMS",  "c170", _C170_IDX["VL_BC_ICMS"]),
    ("ALIQ_ICMS",   "c170", _C170_IDX["ALIQ_ICMS"]),
    ("VL_ICMS",     "c170", _C170_IDX["VL_ICMS"]),
]

HEADERS  = [h for h, _, _ in OUTPUT_COLUMNS]
N_OUTPUT = len(OUTPUT_COLUMNS)

COL_WIDTHS = {
    "NUM_DOC": 14, "COD_ITEM": 16, "DESCR_COMPL": 40, "NCM": 14,
    "CEST": 14, "VL_ITEM": 15, "CST_ICMS": 12, "CFOP": 10,
    "VL_BC_ICMS": 15, "ALIQ_ICMS": 13, "VL_ICMS": 15,
}

# ─── COLUNAS DE SAÍDA — C190 SAÍDAS ─────────────────────────────────────────
# Contexto do C100 pai + todos os campos do C190 (exceto REG)
C190_OUTPUT = [
    ("NUM_DOC",      "c100", _C100_IDX["NUM_DOC"]),
    ("DT_DOC",       "c100", _C100_IDX["DT_DOC"]),
    ("COD_PART",     "c100", _C100_IDX["COD_PART"]),
    ("VL_DOC",       "c100", _C100_IDX["VL_DOC"]),
    ("CST_ICMS",     "c190", _C190_IDX["CST_ICMS"]),
    ("CFOP",         "c190", _C190_IDX["CFOP"]),
    ("ALIQ_ICMS",    "c190", _C190_IDX["ALIQ_ICMS"]),
    ("VL_OPR",       "c190", _C190_IDX["VL_OPR"]),
    ("VL_BC_ICMS",   "c190", _C190_IDX["VL_BC_ICMS"]),
    ("VL_ICMS",      "c190", _C190_IDX["VL_ICMS"]),
    ("VL_BC_ICMS_ST","c190", _C190_IDX["VL_BC_ICMS_ST"]),
    ("VL_ICMS_ST",   "c190", _C190_IDX["VL_ICMS_ST"]),
    ("VL_RED_BC",    "c190", _C190_IDX["VL_RED_BC"]),
    ("VL_IPI",       "c190", _C190_IDX["VL_IPI"]),
    ("COD_OBS",      "c190", _C190_IDX["COD_OBS"]),
]

C190_HEADERS = [h for h, _, _ in C190_OUTPUT]
N_C190_OUT   = len(C190_OUTPUT)

C190_COL_WIDTHS = {
    "NUM_DOC": 14, "DT_DOC": 12, "COD_PART": 18, "VL_DOC": 16,
    "CST_ICMS": 12, "CFOP": 10, "ALIQ_ICMS": 13, "VL_OPR": 16,
    "VL_BC_ICMS": 15, "VL_ICMS": 15, "VL_BC_ICMS_ST": 16,
    "VL_ICMS_ST": 15, "VL_RED_BC": 15, "VL_IPI": 14, "COD_OBS": 14,
}


# ─── PARSER OTIMIZADO ────────────────────────────────────────────────────────
def parse_efd_bytes(raw: bytes) -> dict:
    try:
        text = raw.decode("latin-1")
    except Exception:
        text = raw.decode("utf-8", errors="replace")

    lines = text.splitlines()

    # ── Fase 1: indexar 0200 ─────────────────────────────────────────────
    lookup_0200: dict[str, tuple[str, str]] = {}
    for line in lines:
        if "|0200|" not in line:
            continue
        stripped = line.strip()
        if not stripped:
            continue
        if stripped[0] == "|":
            stripped = stripped[1:]
        if stripped[-1] == "|":
            stripped = stripped[:-1]
        parts = stripped.split("|")
        if parts[0].strip().upper() != "0200":
            continue
        cod  = parts[IDX_0200_COD_ITEM].strip() if len(parts) > IDX_0200_COD_ITEM else ""
        ncm  = parts[IDX_0200_COD_NCM].strip()  if len(parts) > IDX_0200_COD_NCM  else ""
        cest = parts[IDX_0200_CEST].strip()      if len(parts) > IDX_0200_CEST     else ""
        if cod:
            lookup_0200[cod] = (ncm, cest)

    # ── Fase 2: C100 + C170 + C190 ──────────────────────────────────────
    entradas     = []
    saidas       = []
    saidas_c190  = []   # lista de (c100, c190)
    current_c100  = None
    current_c170s = []
    current_c190s = []
    current_oper  = None
    empty_extra   = ("", "")
    cod_item_idx  = _C170_IDX["COD_ITEM"]

    def _flush():
        nonlocal current_c100, current_c170s, current_c190s, current_oper
        if current_c100 is None:
            return
        rec = (current_c100, current_c170s)
        if current_oper == "0":
            entradas.append(rec)
        elif current_oper == "1":
            saidas.append(rec)
            for c190 in current_c190s:
                saidas_c190.append((current_c100, c190))
        current_c100  = None
        current_c170s = []
        current_c190s = []
        current_oper  = None

    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue

        # Filtro rápido: só processa linhas com C1 ou C19
        if "|C1" not in stripped:
            continue

        if stripped[0] == "|":
            stripped = stripped[1:]
        if stripped[-1] == "|":
            stripped = stripped[:-1]

        parts = stripped.split("|")
        reg = parts[0].strip().upper()

        if reg == "C100":
            _flush()
            c100 = parts[:N_C100]
            while len(c100) < N_C100:
                c100.append("")
            current_c100 = c100
            current_oper = parts[1].strip() if len(parts) > 1 else ""

        elif reg == "C170" and current_c100 is not None:
            c170 = parts[:N_C170]
            while len(c170) < N_C170:
                c170.append("")
            cod_item = c170[cod_item_idx].strip()
            extra = lookup_0200.get(cod_item, empty_extra)
            current_c170s.append((c170, extra))

        elif reg == "C190" and current_c100 is not None:
            c190 = parts[:N_C190]
            while len(c190) < N_C190:
                c190.append("")
            current_c190s.append(c190)

    _flush()

    return {
        "entradas":    entradas,
        "saidas":      saidas,
        "saidas_c190": saidas_c190,
        "itens_0200":  len(lookup_0200),
    }


# ─── EXTRAÇÃO DE ARQUIVO ────────────────────────────────────────────────────
def extract_file_from_upload(uploaded) -> bytes | None:
    name = uploaded.name.lower()
    raw  = uploaded.read()

    if name.endswith(".zip"):
        with zipfile.ZipFile(io.BytesIO(raw)) as zf:
            txt_files = [f for f in zf.namelist()
                         if f.lower().endswith(".txt") and not f.startswith("__MACOSX")]
            if not txt_files:
                st.error("Nenhum arquivo .txt encontrado dentro do ZIP.")
                return None
            if len(txt_files) > 1:
                st.info(f"Encontrados {len(txt_files)} .txt no ZIP. Usando: {txt_files[0]}")
            return zf.read(txt_files[0])

    elif name.endswith(".rar"):
        try:
            _install("rarfile")
            import rarfile
            with tempfile.NamedTemporaryFile(suffix=".rar", delete=False) as tmp:
                tmp.write(raw)
                tmp_path = tmp.name
            with rarfile.RarFile(tmp_path) as rf:
                txt_files = [f for f in rf.namelist() if f.lower().endswith(".txt")]
                if not txt_files:
                    st.error("Nenhum arquivo .txt encontrado dentro do RAR.")
                    return None
                if len(txt_files) > 1:
                    st.info(f"Encontrados {len(txt_files)} .txt no RAR. Usando: {txt_files[0]}")
                data = rf.read(txt_files[0])
            os.unlink(tmp_path)
            return data
        except Exception as e:
            st.error(f"Erro ao extrair RAR: {e}")
            return None
    else:
        return raw


# ─── GERADOR XLSX ────────────────────────────────────────────────────────────
def build_xlsx(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Estilos compartilhados
    thin       = Side(style="thin", color="999999")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    align_c    = Alignment(horizontal="center", vertical="center")
    align_l    = Alignment(horizontal="left",   vertical="center")
    font_d     = Font(name="Arial", size=9)
    fill_z     = PatternFill("solid", fgColor="F2F7FB")
    fill_w     = PatternFill("solid", fgColor="FFFFFF")

    # Cabeçalhos por aba
    FILL_HEADER = {
        "ENTRADAS":    PatternFill("solid", fgColor="1F4E79"),
        "SAÍDAS":      PatternFill("solid", fgColor="1F4E79"),
        "C190 SAÍDAS": PatternFill("solid", fgColor="6B21A8"),
    }
    font_h = Font(name="Arial", bold=True, color="FFFFFF", size=10)

    wb = Workbook()
    wb.remove(wb.active)

    # ── Helper: aplica estilo numa célula ────────────────────────────────
    def _style(cell, font, fill, align):
        cell.font, cell.fill, cell.alignment, cell.border = font, fill, align, border_all

    # ── Aba ENTRADAS / SAÍDAS (C100+C170+0200) ──────────────────────────
    def _extract_row(c100, c170, extra):
        row = []
        for _, origem, idx in OUTPUT_COLUMNS:
            if origem == "c100":
                row.append(c100[idx])
            elif origem == "c170":
                row.append(c170[idx])
            else:
                row.append(extra[idx])
        return row

    def write_c170_sheet(title: str, records: list):
        ws      = wb.create_sheet(title=title)
        fill_hd = FILL_HEADER[title]

        # Cabeçalho
        for col_idx, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            _style(cell, font_h, fill_hd, align_c)

        # Dados
        row_num = 1
        for c100, c170s in records:
            if not c170s:
                continue
            for c170, extra in c170s:
                row_num += 1
                fill = fill_z if row_num % 2 == 0 else fill_w
                vals = _extract_row(c100, c170, extra)
                for col_idx, val in enumerate(vals, 1):
                    cell = ws.cell(row=row_num, column=col_idx, value=val)
                    al = align_l if HEADERS[col_idx - 1] == "DESCR_COMPL" else align_c
                    _style(cell, font_d, fill, al)

        if row_num == 1:
            cell = ws.cell(row=2, column=1, value=f"Nenhum registro de {title} encontrado.")
            cell.font = Font(name="Arial", bold=True, size=11, color="CC0000")

        # Larguras
        for col_idx, h in enumerate(HEADERS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(h, 12)

        # Congelar cabeçalho + autofiltro
        ws.freeze_panes = "A2"
        if row_num > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(N_OUTPUT)}{row_num}"

    # ── Aba C190 SAÍDAS ─────────────────────────────────────────────────
    def write_c190_sheet(saidas_c190: list):
        title   = "C190 SAÍDAS"
        ws      = wb.create_sheet(title=title)
        fill_hd = FILL_HEADER[title]

        # Cabeçalho
        for col_idx, h in enumerate(C190_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            _style(cell, font_h, fill_hd, align_c)

        # Dados
        row_num = 1
        for c100, c190 in saidas_c190:
            row_num += 1
            fill = fill_z if row_num % 2 == 0 else fill_w
            for col_idx, (_, origem, idx) in enumerate(C190_OUTPUT, 1):
                if origem == "c100":
                    val = c100[idx]
                else:
                    val = c190[idx]
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                _style(cell, font_d, fill, align_c)

        if row_num == 1:
            cell = ws.cell(row=2, column=1, value="Nenhum registro C190 de saída encontrado.")
            cell.font = Font(name="Arial", bold=True, size=11, color="CC0000")

        # Larguras
        for col_idx, h in enumerate(C190_HEADERS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = C190_COL_WIDTHS.get(h, 12)

        # Congelar cabeçalho + autofiltro
        ws.freeze_panes = "A2"
        if row_num > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(N_C190_OUT)}{row_num}"

    # ── Gera as 3 abas ──────────────────────────────────────────────────
    write_c170_sheet("ENTRADAS", data["entradas"])
    write_c170_sheet("SAÍDAS",   data["saidas"])
    write_c190_sheet(data["saidas_c190"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── UTILITÁRIOS ─────────────────────────────────────────────────────────────
def detect_efd_type(raw: bytes) -> str:
    try:
        text = raw.decode("latin-1")
    except Exception:
        text = raw.decode("utf-8", errors="replace")
    for line in text.splitlines():
        if "|0000|" not in line:
            continue
        parts = line.split("|")
        for p in parts:
            if p.strip() in ("007","008","009","010","011","012","013",
                              "014","015","016","017","018","019","020"):
                return "EFD ICMS/IPI"
        return "EFD Contribuições"
    return "Não identificado"


def count_records(raw: bytes) -> dict:
    try:
        text = raw.decode("latin-1")
    except Exception:
        text = raw.decode("utf-8", errors="replace")
    c100 = c170 = c190 = r0200 = 0
    for line in text.splitlines():
        s = line.strip()
        if   s.startswith("|C100|"): c100  += 1
        elif s.startswith("|C170|"): c170  += 1
        elif s.startswith("|C190|"): c190  += 1
        elif s.startswith("|0200|"): r0200 += 1
    return {"C100": c100, "C170": c170, "C190": c190, "0200": r0200}


# ─── UI STREAMLIT ────────────────────────────────────────────────────────────
def main():
    st.set_page_config(page_title="EFD Extrator C100+C170+C190", page_icon="📊", layout="wide")

    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
    .main .block-container { max-width: 960px; padding-top: 2rem; }
    .hero-title {
        font-family: 'Inter', sans-serif; font-size: 2rem; font-weight: 700;
        color: #1F4E79; text-align: center; margin-bottom: 0.2rem;
    }
    .hero-sub {
        font-family: 'Inter', sans-serif; font-size: 1rem;
        color: #666; text-align: center; margin-bottom: 2rem;
    }
    .stat-card {
        background: linear-gradient(135deg, #f8fafc, #eef2f7);
        border: 1px solid #e2e8f0; border-radius: 12px;
        padding: 1.2rem; text-align: center;
    }
    .stat-card h3 {
        font-family: 'Inter', sans-serif; font-size: 1.6rem;
        font-weight: 700; color: #1F4E79; margin: 0;
    }
    .stat-card p {
        font-family: 'Inter', sans-serif; font-size: 0.85rem;
        color: #64748b; margin: 0.3rem 0 0 0;
    }
    .info-badge {
        display: inline-block; background: #e0f2fe; color: #0369a1;
        font-family: 'Inter', sans-serif; font-size: 0.8rem; font-weight: 600;
        padding: 0.3rem 0.8rem; border-radius: 20px; margin: 0.2rem;
    }
    div[data-testid="stFileUploader"] {
        border: 2px dashed #2E75B6 !important;
        border-radius: 12px !important; padding: 1rem !important;
    }
    .stDownloadButton > button {
        background: linear-gradient(135deg, #1F4E79, #2E75B6) !important;
        color: white !important; font-weight: 600 !important;
        border-radius: 10px !important; padding: 0.7rem 2rem !important;
        border: none !important; width: 100% !important; font-size: 1rem !important;
    }
    .stDownloadButton > button:hover {
        background: linear-gradient(135deg, #163a5c, #1F4E79) !important;
    }
    footer {visibility: hidden;}
    </style>
    """, unsafe_allow_html=True)

    st.markdown('<div class="hero-title">📊 EFD Extrator C100 + C170 + C190</div>', unsafe_allow_html=True)
    st.markdown('<div class="hero-sub">Extraia registros de Notas Fiscais do seu SPED em segundos</div>', unsafe_allow_html=True)

    st.markdown("""
    <div style="text-align:center; margin-bottom: 1.5rem;">
        <span class="info-badge">EFD ICMS/IPI</span>
        <span class="info-badge">EFD Contribuições</span>
        <span class="info-badge">TXT • ZIP • RAR</span>
        <span class="info-badge">NCM + CEST (0200)</span>
        <span class="info-badge">C190 Saídas</span>
    </div>
    """, unsafe_allow_html=True)

    st.divider()

    uploaded = st.file_uploader(
        "📂 Importe o arquivo EFD",
        type=["txt", "zip", "rar"],
        help="Aceita arquivos .txt, .zip ou .rar contendo o arquivo EFD"
    )

    if uploaded is not None:
        with st.spinner("Extraindo arquivo..."):
            raw = extract_file_from_upload(uploaded)

        if raw is None:
            st.stop()

        efd_type     = detect_efd_type(raw)
        counts       = count_records(raw)
        file_size_mb = len(raw) / (1024 * 1024)
        tipo_label   = "ICMS/IPI" if "/" in efd_type else efd_type.split()[-1]

        c1, c2, c3, c4, c5, c6 = st.columns(6)
        with c1:
            st.markdown(f'<div class="stat-card"><h3>{tipo_label}</h3><p>Tipo EFD</p></div>', unsafe_allow_html=True)
        with c2:
            st.markdown(f'<div class="stat-card"><h3>{counts["C100"]:,}</h3><p>Registros C100</p></div>', unsafe_allow_html=True)
        with c3:
            st.markdown(f'<div class="stat-card"><h3>{counts["C170"]:,}</h3><p>Registros C170</p></div>', unsafe_allow_html=True)
        with c4:
            st.markdown(f'<div class="stat-card"><h3>{counts["C190"]:,}</h3><p>Registros C190</p></div>', unsafe_allow_html=True)
        with c5:
            st.markdown(f'<div class="stat-card"><h3>{counts["0200"]:,}</h3><p>Itens (0200)</p></div>', unsafe_allow_html=True)
        with c6:
            st.markdown(f'<div class="stat-card"><h3>{file_size_mb:.1f} MB</h3><p>Tamanho</p></div>', unsafe_allow_html=True)

        if counts["C100"] == 0:
            st.warning("⚠️ Nenhum registro C100 encontrado. Verifique se é um arquivo EFD válido.")
            st.stop()

        st.divider()

        if st.button("⚡ Processar e Gerar Planilha", use_container_width=True, type="primary"):
            t0 = time.time()

            with st.spinner("Processando registros..."):
                data = parse_efd_bytes(raw)

            n_ent       = len(data["entradas"])
            n_sai       = len(data["saidas"])
            n_itens_ent = sum(len(c170s) for _, c170s in data["entradas"])
            n_itens_sai = sum(len(c170s) for _, c170s in data["saidas"])
            n_c190_sai  = len(data["saidas_c190"])

            with st.spinner("Gerando planilha XLSX..."):
                xlsx_bytes = build_xlsx(data)

            elapsed = time.time() - t0

            st.success(f"✅ Processado em {elapsed:.1f}s — {data['itens_0200']:,} itens no cadastro 0200")

            col_e, col_s, col_c = st.columns(3)
            with col_e:
                st.markdown(f"""
                <div class="stat-card" style="border-left: 4px solid #16a34a;">
                    <h3 style="color:#16a34a;">{n_ent}</h3>
                    <p>Notas de Entrada</p>
                    <p style="font-size:0.75rem; color:#94a3b8;">{n_itens_ent:,} itens (C170)</p>
                </div>""", unsafe_allow_html=True)
            with col_s:
                st.markdown(f"""
                <div class="stat-card" style="border-left: 4px solid #dc2626;">
                    <h3 style="color:#dc2626;">{n_sai}</h3>
                    <p>Notas de Saída</p>
                    <p style="font-size:0.75rem; color:#94a3b8;">{n_itens_sai:,} itens (C170)</p>
                </div>""", unsafe_allow_html=True)
            with col_c:
                st.markdown(f"""
                <div class="stat-card" style="border-left: 4px solid #6B21A8;">
                    <h3 style="color:#6B21A8;">{n_c190_sai}</h3>
                    <p>C190 Saídas</p>
                    <p style="font-size:0.75rem; color:#94a3b8;">Consolidação ICMS</p>
                </div>""", unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            base_name   = Path(uploaded.name).stem
            output_name = f"{base_name}_C100_C170_C190.xlsx"

            st.download_button(
                label=f"📥 Baixar {output_name}",
                data=xlsx_bytes,
                file_name=output_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    st.markdown("<br>", unsafe_allow_html=True)
    st.divider()
    st.markdown("""
    <div style="text-align:center; padding: 0.5rem 0;">
        <span style="font-family:'Inter',sans-serif; font-size:0.75rem; color:#94a3b8;">
            UFISCAL — Inteligência em Negócios • Layout conforme Guia Prático EFD v3.1.8 / EFD Contribuições v1.35
        </span>
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
